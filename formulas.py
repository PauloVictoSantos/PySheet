from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = "estoque.xlsx"

workbook = load_workbook(file_path)

sheet = workbook["Estoque"]

col_nome_produto = 1
col_valor_fornecedor = 2
col_lucratividade = 3
col_quantidade = 4

col_preco_venda = col_quantidade + 1
col_lucro_total = col_preco_venda + 1
col_valor_total = col_lucro_total + 1

sheet.cell(row=1, column=col_preco_venda, value="Preço de Venda")
sheet.cell(row=1, column=col_lucro_total, value="Lucro Total")
sheet.cell(row=1, column=col_valor_total, value="Valor Total")

max_row = sheet.max_row

for row in range(2, max_row + 1):
    cell_valor_fornecedor = sheet.cell(row=row, column=col_valor_fornecedor)
    cell_lucratividade = sheet.cell(row=row, column=col_lucratividade)
    cell_quantidade = sheet.cell(row=row, column=col_quantidade)
    cell_preco_venda = sheet.cell(row=row, column=col_preco_venda)

    formula_preco_venda = (
        f"={cell_valor_fornecedor.coordinate}*(1+{cell_lucratividade.coordinate}/100)"
    )
    cell_preco_venda.value = formula_preco_venda

    formula_lucro_total = f"=({cell_preco_venda.coordinate}-{cell_lucratividade.coordinate})*{cell_quantidade.coordinate}"
    sheet.cell(row=row, column=col_lucro_total).value = formula_lucro_total

    formula_valor_total = f"={cell_preco_venda.coordinate}*{cell_quantidade.coordinate}"
    sheet.cell(row=row, column=col_valor_total).value = formula_valor_total

linha_total = max_row + 2
sheet.cell(row=linha_total, column=col_nome_produto, value="Totais Gerais")
sheet.merge_cells(
    start_row=linha_total,
    start_column=col_nome_produto,
    end_row=linha_total,
    end_column=col_quantidade,
)

formula_total_lucro = f"=SUM({get_column_letter(col_lucro_total)}2:{get_column_letter(col_lucro_total)}{max_row})"
sheet.cell(row=linha_total, column=col_lucro_total).value = formula_lucro_total

formula_total_valor = f"=SUM({get_column_letter(col_valor_total)}2:{get_column_letter(col_valor_total)}{max_row})"
sheet.cell(row=linha_total, column=col_valor_total).value = formula_total_valor


workbook.save(file_path)