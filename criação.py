import random
from openpyxl import Workbook

workbook = Workbook()

sheet = workbook.active
sheet.title = "Estoque"


headers = ["Nome do produto", "Valor do fornecedor", "Lucratividade (%)", "Quatidade"]

for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num, value=header)


def gerar_nome_produto():
    prefixos = ["Super", "Mega", "Ultra", "Power", "Max"]
    tipos = ["Widget", "Gadget", "Device", "Intrument", "Appliance"]
    sufixos = ["Plus", "Pro", "X", "2000", "Elite"]
    return f"{random.choice(prefixos)} {random.choice(tipos)} {random.choice(sufixos)}"


num_produtos = 50

for row_nm in range(2, num_produtos + 2):
    nome_prodto = gerar_nome_produto()
    valor_fornecedor = round(random.uniform(10.0, 500.0), 2)
    lucratividade = random.randint(10, 100)
    quantidade = random.randint(1, 100)

    sheet.cell(row=row_nm, column=1, value=nome_prodto)
    sheet.cell(row=row_nm, column=2, value=valor_fornecedor)
    sheet.cell(row=row_nm, column=3, value=lucratividade)
    sheet.cell(row=row_nm, column=4, value=quantidade)


file_path = 'estoque.xlsx'
workbook.save(file_path)